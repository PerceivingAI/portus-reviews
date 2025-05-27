# portus_sa_module/sa_writer.py

from pathlib import Path
from openpyxl import load_workbook
from portus_sa_module.sa_engine import score_individual_review


def write_individual_scores(clean_file_path: Path) -> None:
    if not clean_file_path.exists():
        raise FileNotFoundError(clean_file_path)

    wb = load_workbook(clean_file_path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    headers_lc = [str(h).lower() if h else "" for h in headers]

    if "sa_score" in headers_lc:
        sa_col = headers_lc.index("sa_score") + 1
    else:
        sa_col = len(headers) + 1
        ws.cell(row=1, column=sa_col, value="sa_score")

    title_idx = next((i for i, h in enumerate(headers_lc) if h in ("title", "reviewtitle")), None)
    text_idx  = headers_lc.index("text") if "text" in headers_lc else None
    if text_idx is None:
        raise ValueError("'text' column missing")

    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = row[title_idx] if title_idx is not None else ""
        text  = row[text_idx]  if text_idx  is not None else ""
        if not (str(title).strip() or str(text).strip()):
            continue
        score = score_individual_review(str(title), str(text))
        ws.cell(row=r, column=sa_col, value=score)

    wb.save(clean_file_path)
    wb.close()
