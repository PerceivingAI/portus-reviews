# reply_module\reply_writer.py

from pathlib import Path
from openpyxl import load_workbook

def write_replies_to_excel(clean_file_path: Path, replies: list[str], reply_column: str):
    """
    Writes generated replies into the Excel file at the specified reply column.
    
    Args:
        clean_file_path (Path): Path to the Excel file.
        replies (list[str]): List of generated replies.
        reply_column (str): The header name of the column to write replies to.
    """
    if not clean_file_path.exists():
        print(f"❌ File not found: {clean_file_path}")
        return

    wb = load_workbook(clean_file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if reply_column not in headers:
        #print(f"❌ Reply column '{reply_column}' not found in sheet headers.")
        wb.close()
        return

    reply_col_index = headers.index(reply_column) + 1

    for idx, reply in enumerate(replies, start=2):
        try:
            ws.cell(row=idx, column=reply_col_index, value=reply)
            #print(f"✅ Wrote reply to row {idx}")
        except Exception as e:
            print(f"❌ Failed to write reply at row {idx}: {e}")

    try:
        wb.save(clean_file_path)
        #print("✅ Workbook saved.")
    except Exception as e:
        print(f"❌ Failed to save workbook: {e}")
    wb.close()
