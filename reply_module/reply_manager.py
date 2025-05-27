# reply_module/reply_manager.py

from pathlib import Path
from openpyxl import load_workbook

from reply_module.reply_engine import generate_reply

PROVIDER_REPLY_COLUMNS = {
    "openai": "Reply OAI",
    "google": "Reply Go",
    "xai":    "Reply XAI",
}

SITE_REVIEW_COLUMNS = {
    "tripadvisor": "text",
    "google":      "text",
    "booking":     "text",
    "expedia":     "text",
}

SITE_TITLE_COLUMNS = {
    "tripadvisor": "title",
    "google":      "title",
    "booking":     "reviewTitle",
    "expedia":     "title",
}


def handle_reply_generation(provider: str, active_sites: list[str], clean_file_path: Path) -> None:
    print("🧪 Provider:", provider)
    print("🧪 Active review sites:", active_sites)
    print("🧪 Clean file path:", clean_file_path)

    reply_column = PROVIDER_REPLY_COLUMNS.get(provider.lower())
    if not reply_column:
        print(f"❌ Unsupported provider: {provider}")
        return

    if not clean_file_path.exists():
        print(f"❌ Clean file not found: {clean_file_path}")
        return

    wb = load_workbook(clean_file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]

    if reply_column in headers:
        reply_col_idx = headers.index(reply_column)
    else:
        reply_col_idx = len(headers)
        ws.cell(row=1, column=reply_col_idx + 1, value=reply_column)
        #print(f"➕ Added reply column: '{reply_column}'")

    site = active_sites[0]
    review_col = SITE_REVIEW_COLUMNS[site]
    title_col  = SITE_TITLE_COLUMNS[site]

    if review_col.lower() not in headers_lower:
        #print(f"❌ Review column '{review_col}' not found in sheet.")
        wb.close()
        return

    review_idx = headers_lower.index(review_col.lower())
    title_idx = headers_lower.index(title_col.lower()) if title_col.lower() in headers_lower else None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        existing_reply = row[reply_col_idx] if reply_col_idx < len(row) else None
        review_text    = row[review_idx] if review_idx < len(row) else ""
        title_text     = row[title_idx] if title_idx is not None and title_idx < len(row) else ""

        if existing_reply or not review_text:
            continue

        reply = generate_reply(provider, review_text, title_text)
        ws.cell(row=row_idx, column=reply_col_idx + 1, value=reply)
        #print(f"💾 Wrote reply to row {row_idx}")

    wb.save(clean_file_path)
    wb.close()
    #print("✅ Workbook saved.")
