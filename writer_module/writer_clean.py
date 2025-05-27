# writer_module\writer_clean.py

from pathlib import Path
from openpyxl import load_workbook, Workbook

def write_clean_excel(site: str, input_file: Path, output_file: Path, columns: list):
    wb_in = load_workbook(input_file)
    ws_in = wb_in.active

    headers = [cell.value for cell in ws_in[1]]
    col_indices = [headers.index(c) for c in columns if c in headers]

    synthesize_text = False
    liked_idx = disliked_idx = None
    if site == "booking":
        try:
            liked_idx = headers.index("likedText")
            disliked_idx = headers.index("dislikedText")
            synthesize_text = True
        except ValueError:
            synthesize_text = False

    new_headers = [headers[i] for i in col_indices]
    if synthesize_text and "text" not in new_headers:
        new_headers.append("text")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append(new_headers)

    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=2):
        clean_row = [row[i] for i in col_indices]

        if site == "booking" and synthesize_text:
            liked = row[liked_idx] or ""
            disliked = row[disliked_idx] or ""
            combined_text = f"{liked.strip()} {disliked.strip()}".strip()
            clean_row.append(combined_text)
            text_val = combined_text
        else:
            text_val = None
            if "text" in columns:
                try:
                    text_idx = headers.index("text")
                    text_val = row[text_idx]
                except ValueError:
                    text_val = ""

        title_val = ""
        for tkey in ("title", "reviewTitle"):
            if tkey in headers:
                try:
                    idx = headers.index(tkey)
                    title_val = row[idx]
                    break
                except ValueError:
                    continue

        if not title_val and not text_val:
            continue

        ws_out.append(clean_row)

    wb_out.save(output_file)
